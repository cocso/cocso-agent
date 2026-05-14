"""cocso_settlement — COCSO 표준 의약품 정산 수수료 내역서 생성 도구.

번들 템플릿:
  plugins/cocso_settlement/template.xlsx  (시트: "의약품 정산 수수료 내역서")
  - 식별자 T1: COCSO_SETTLEMENT_COMMISSION_SHEET_V1
  - 데이터 영역: B4:L5000 (편집 가능)
  - 통계 영역: N3:R5000 (수식 자동 집계 — 건드리지 X)
  - 회사명: L2

워크플로우:
  1. 템플릿 복사 → 출력 경로
  2. L2 = company_name
  3. items[] 의 각 dict → B{row}:L{row} 셀 채우기 (row=4부터)
  4. 저장 (수식 자동 재계산은 Excel 열 때 일어남)

Tools:
  cocso_settlement_create        — 정산 내역서 신규 생성
  cocso_settlement_template_info — 템플릿 컬럼·키 매핑 조회
"""
from __future__ import annotations

import json
import logging
import os
import re
import shutil
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

try:
    import openpyxl  # type: ignore[import-untyped]
    _OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]
    _OPENPYXL_AVAILABLE = False


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

PLUGIN_DIR = Path(__file__).parent
TEMPLATE_PATH = PLUGIN_DIR / "template.xlsx"
SHEET_NAME = "의약품 정산 수수료 내역서"
COMPANY_CELL = "L2"
DATA_START_ROW = 4
DATA_MAX_ROW = 5000  # 템플릿 통계 수식의 SUM 범위 상한

# 컬럼 정의 — order matters (B, C, D, ..., L)
# 각 항목: (column_letter, korean_label, snake_key, expected_type)
COLUMNS = [
    ("B", "정산코드",                "settlement_code",       "str"),
    ("C", "처방 병원명",             "hospital_name",         "str"),
    ("D", "처방 병원 사업자번호",    "hospital_biz_id",       "str"),
    ("E", "제약사명",                "manufacturer",          "str"),
    ("F", "보험코드",                "insurance_code",        "str"),
    ("G", "제품명",                  "product_name",          "str"),
    ("H", "단가(원)",                "unit_price",            "number"),
    ("I", "수량",                    "quantity",              "number"),
    ("J", "처방금액(원)",            "prescription_amount",   "number"),
    ("K", "정산금액(원, VAT 포함)",  "settlement_amount",     "number"),
    ("L", "비고",                    "note",                  "str"),
]

# 빠른 조회용 — 한국어 + snake_key 모두 매칭 가능
_KEY_LOOKUP = {}
for col, kor, snake, _t in COLUMNS:
    _KEY_LOOKUP[snake] = col
    _KEY_LOOKUP[kor] = col
    # 공백 / 괄호 / 줄바꿈 차이 흡수
    _KEY_LOOKUP[kor.replace(" ", "")] = col
    _KEY_LOOKUP[kor.replace(",", "").replace("(", "").replace(")", "")] = col


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _err(msg: str, **extra) -> str:
    return json.dumps({"error": msg, **extra}, ensure_ascii=False)


def _ok(**fields) -> str:
    return json.dumps({"ok": True, **fields}, ensure_ascii=False, default=str)


def _resolve(path: str) -> Path:
    return Path(os.path.expanduser(path)).resolve()


def _check_sdk() -> Optional[str]:
    if not _OPENPYXL_AVAILABLE:
        return _err(
            "openpyxl not installed",
            install="pip install 'openpyxl>=3.1,<4'",
        )
    if not TEMPLATE_PATH.exists():
        return _err(
            "template missing",
            expected_path=str(TEMPLATE_PATH),
            hint="Plugin install incomplete — reinstall cocso-agent.",
        )
    return None


def _normalize_item(item: Dict[str, Any]) -> Dict[str, Any]:
    """Map an item's keys (Korean OR snake_case OR mixed) to column letters.

    Returns ``{column_letter: value}`` for cells that should be written.
    Unknown keys are dropped silently (caller can introspect via
    ``cocso_settlement_template_info``).
    """
    if not isinstance(item, dict):
        return {}
    out: Dict[str, Any] = {}
    for k, v in item.items():
        col = _KEY_LOOKUP.get(k) or _KEY_LOOKUP.get(str(k).strip())
        if col:
            out[col] = v
    return out


def _coerce_number(value: Any) -> Any:
    """Best-effort numeric coercion for amount/qty fields."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    s = str(value).replace(",", "").strip()
    if not s:
        return None
    try:
        if "." in s:
            return float(s)
        return int(s)
    except ValueError:
        return value  # leave as-is; user inspects via cell value


# ---------------------------------------------------------------------------
# Tool: cocso_settlement_create
# ---------------------------------------------------------------------------

def cocso_settlement_create(args: Dict[str, Any], **_kw) -> str:
    """Create a new settlement workbook from the bundled template."""
    if (e := _check_sdk()) is not None:
        return e

    output_path = (args.get("output_path") or "").strip()
    company_name = (args.get("company_name") or "").strip()
    items = args.get("items") or []
    overwrite = bool(args.get("overwrite", False))

    if not output_path:
        return _err("output_path required (e.g. ~/Documents/정산_2026-05.xlsx)")
    if not isinstance(items, list):
        return _err("items must be a list of objects")

    dst = _resolve(output_path)
    if not str(dst).lower().endswith(".xlsx"):
        return _err("output_path must end with .xlsx", output_path=str(dst))
    if dst.exists() and not overwrite:
        return _err(
            "output already exists. Pass overwrite=true to replace.",
            output_path=str(dst),
        )

    if len(items) > (DATA_MAX_ROW - DATA_START_ROW + 1):
        return _err(
            "items exceed template capacity",
            max_rows=DATA_MAX_ROW - DATA_START_ROW + 1,
            received=len(items),
        )

    try:
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(TEMPLATE_PATH, dst)

        wb = openpyxl.load_workbook(dst)
        if SHEET_NAME not in wb.sheetnames:
            wb.close()
            return _err(
                f"template missing sheet '{SHEET_NAME}'",
                available=wb.sheetnames,
            )
        ws = wb[SHEET_NAME]

        if company_name:
            ws[COMPANY_CELL] = company_name

        rows_written = 0
        unknown_keys: set = set()
        for idx, raw in enumerate(items):
            row = DATA_START_ROW + idx
            mapped = _normalize_item(raw)
            if not mapped and isinstance(raw, dict):
                unknown_keys.update(str(k) for k in raw.keys())
                continue
            for col, val in mapped.items():
                # 숫자 타입 컬럼 (H/I/J/K) 은 coerce
                if col in ("H", "I", "J", "K"):
                    val = _coerce_number(val)
                ws[f"{col}{row}"] = val
            rows_written += 1

        wb.save(dst)
        wb.close()

        return _ok(
            output_path=str(dst),
            company_name=company_name or None,
            rows_written=rows_written,
            items_received=len(items),
            unknown_keys=sorted(unknown_keys) if unknown_keys else None,
            note=(
                "Excel에서 파일을 열면 통계 영역 (N3:R) 의 수식이 자동 재계산됩니다. "
                "openpyxl만으로 미리 계산된 값은 보이지 않을 수 있습니다."
            ),
        )
    except Exception as exc:
        return _err(f"create failed: {exc}", output_path=str(dst))


# ---------------------------------------------------------------------------
# Tool: cocso_settlement_template_info
# ---------------------------------------------------------------------------

def cocso_settlement_template_info(args: Dict[str, Any], **_kw) -> str:
    """Return template metadata + accepted column key mappings."""
    if (e := _check_sdk()) is not None:
        return e
    columns = [
        {
            "column": col,
            "korean_label": kor,
            "snake_key": snake,
            "type": typ,
        }
        for col, kor, snake, typ in COLUMNS
    ]
    return _ok(
        template_path=str(TEMPLATE_PATH),
        sheet_name=SHEET_NAME,
        company_cell=COMPANY_CELL,
        data_start_row=DATA_START_ROW,
        data_max_row=DATA_MAX_ROW,
        max_rows=DATA_MAX_ROW - DATA_START_ROW + 1,
        columns=columns,
        accepted_key_forms=(
            "각 항목 dict의 키는 한국어 라벨('정산코드') 또는 snake_case "
            "('settlement_code') 둘 다 인식됩니다. 모르는 키는 조용히 무시 — "
            "응답의 unknown_keys 필드로 확인 가능."
        ),
    )


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------

CREATE_SCHEMA = {
    "name": "cocso_settlement_create",
    "description": (
        "COCSO 표준 의약품 정산 수수료 내역서를 새로 생성합니다. "
        "번들 템플릿을 복사해 회사명과 정산 항목 row들을 채워 저장. "
        "통계 영역(N3:R)은 템플릿의 수식이 자동 집계하므로 별도 처리 "
        "필요 없음. 사용자가 의약품 정산서·수수료 내역 작성을 요청할 "
        "때 사용. 항목 dict 키는 한국어('정산코드') 또는 snake_case "
        "('settlement_code') 둘 다 가능."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "output_path": {
                "type": "string",
                "description": "저장 경로 (.xlsx). ~ 확장 지원.",
            },
            "company_name": {
                "type": "string",
                "description": "L2 셀에 들어갈 회사명. 비워두면 템플릿 기본값 유지.",
            },
            "items": {
                "type": "array",
                "description": (
                    "정산 항목 배열. 각 항목은 dict — 키는 한국어 라벨 "
                    "(정산코드, 처방 병원명, 처방 병원 사업자번호, 제약사명, "
                    "보험코드, 제품명, 단가(원), 수량, 처방금액(원), "
                    "정산금액(원, VAT 포함), 비고) 또는 snake_case "
                    "(settlement_code, hospital_name, hospital_biz_id, "
                    "manufacturer, insurance_code, product_name, unit_price, "
                    "quantity, prescription_amount, settlement_amount, note) "
                    "둘 다 가능."
                ),
                "items": {"type": "object"},
            },
            "overwrite": {
                "type": "boolean",
                "description": "출력 파일이 이미 있으면 덮어쓸지. default false.",
            },
        },
        "required": ["output_path", "items"],
    },
}

INFO_SCHEMA = {
    "name": "cocso_settlement_template_info",
    "description": (
        "정산 내역서 템플릿 메타데이터 조회 — 컬럼 순서, 한국어 라벨, "
        "snake_case 키, 타입, 데이터 영역 범위. cocso_settlement_create "
        "호출 전 키 매핑을 확인할 때 사용."
    ),
    "parameters": {"type": "object", "properties": {}},
}


# ---------------------------------------------------------------------------
# Tool: cocso_settlement_validate
# ---------------------------------------------------------------------------

def cocso_settlement_validate(args: Dict[str, Any], **_kw) -> str:
    """Items 배열을 cocso_settlement_create 호출 전에 검증.

    감지 항목 (각 항목 row 단위):
      - 처방금액 ≠ 단가 × 수량 (오차 1원 이내 허용)
      - 정산금액 < 처방금액 (보통 정산 = 처방 + VAT 인데 작으면 의심)
      - 핵심 컬럼 빠짐 (정산코드 / 처방 병원명 / 제품명 / 단가 / 수량)
      - 사업자번호 형식 (3-2-5 자리 또는 10자리)
      - 음수 / 0 단가·수량
      - 같은 (정산코드, 제품명) 중복 row

    응답: ``{ok, item_count, errors[], warnings[], summary}``.
    errors > 0 이면 cocso_settlement_create 호출 전 사용자 확인 필수.
    """
    if (e := _check_sdk()) is not None:
        return e
    items = args.get("items") or []
    if not isinstance(items, list):
        return _err("items must be a list of objects")

    errors: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []
    seen_keys: Dict[tuple, int] = {}
    biz_pattern = re.compile(r"^\d{3}-?\d{2}-?\d{5}$")
    total_amount = 0.0
    total_settlement = 0.0

    for idx, item in enumerate(items):
        if not isinstance(item, dict):
            errors.append({"row": idx + 1, "field": "_root",
                           "msg": "item is not a dict"})
            continue

        # Helper: pull value by korean OR snake key
        def _get(*keys):
            for k in keys:
                if k in item and item[k] not in (None, ""):
                    return item[k]
            return None

        code = _get("정산코드", "settlement_code")
        hospital = _get("처방 병원명", "hospital_name")
        product = _get("제품명", "product_name")
        biz_id = _get("처방 병원 사업자번호", "hospital_biz_id")
        unit = _get("단가(원)", "unit_price")
        qty = _get("수량", "quantity")
        prescription = _get("처방금액(원)", "prescription_amount")
        settlement = _get("정산금액(원, VAT 포함)", "settlement_amount")

        # 핵심 컬럼 누락
        for label, val in (("정산코드", code), ("처방 병원명", hospital),
                           ("제품명", product), ("단가", unit), ("수량", qty)):
            if val in (None, ""):
                errors.append({"row": idx + 1, "field": label,
                               "msg": f"{label} 값 비어있음"})

        # 숫자 변환
        unit_n = _coerce_number(unit) if unit is not None else None
        qty_n = _coerce_number(qty) if qty is not None else None
        pres_n = _coerce_number(prescription) if prescription is not None else None
        sett_n = _coerce_number(settlement) if settlement is not None else None

        # 처방금액 ≠ 단가 × 수량
        if isinstance(unit_n, (int, float)) and isinstance(qty_n, (int, float)) \
                and isinstance(pres_n, (int, float)):
            expected = unit_n * qty_n
            if abs(pres_n - expected) > 1:
                errors.append({
                    "row": idx + 1, "field": "처방금액",
                    "msg": f"처방금액 {pres_n:.0f} ≠ 단가 {unit_n} × 수량 {qty_n} = {expected:.0f}",
                })

        # 정산금액 < 처방금액
        if isinstance(pres_n, (int, float)) and isinstance(sett_n, (int, float)):
            if sett_n < pres_n - 1:
                warnings.append({
                    "row": idx + 1, "field": "정산금액",
                    "msg": f"정산금액 {sett_n:.0f} < 처방금액 {pres_n:.0f} (보통 정산 ≥ 처방 + VAT)",
                })

        # 음수·0 단가/수량
        for label, val in (("단가", unit_n), ("수량", qty_n)):
            if isinstance(val, (int, float)) and val <= 0:
                warnings.append({"row": idx + 1, "field": label,
                                 "msg": f"{label}이 0 또는 음수: {val}"})

        # 사업자번호 형식
        if biz_id and not biz_pattern.match(str(biz_id).strip()):
            warnings.append({"row": idx + 1, "field": "처방 병원 사업자번호",
                             "msg": f"형식 이상: {biz_id} (10자리 또는 NNN-NN-NNNNN)"})

        # 중복 (정산코드, 제품명)
        if code and product:
            key = (str(code), str(product))
            if key in seen_keys:
                warnings.append({
                    "row": idx + 1, "field": "_duplicate",
                    "msg": f"row {seen_keys[key]} 와 같은 (정산코드={code}, 제품명={product}) 중복",
                })
            else:
                seen_keys[key] = idx + 1

        # 합계
        if isinstance(pres_n, (int, float)):
            total_amount += pres_n
        if isinstance(sett_n, (int, float)):
            total_settlement += sett_n

    return _ok(
        item_count=len(items),
        errors=errors,
        warnings=warnings,
        error_count=len(errors),
        warning_count=len(warnings),
        summary={
            "total_prescription": round(total_amount, 2),
            "total_settlement": round(total_settlement, 2),
            "has_blockers": len(errors) > 0,
        },
        note=(
            "errors > 0 이면 cocso_settlement_create 호출 전 사용자에게 알리고 "
            "수정 의향 확인. warnings 만 있으면 진행하되 사용자에게 요약."
        ),
    )


VALIDATE_SCHEMA = {
    "name": "cocso_settlement_validate",
    "description": (
        "정산 items 배열을 cocso_settlement_create 호출 전에 자동 검증. "
        "처방금액 = 단가×수량 / 정산금액 ≥ 처방금액 / 핵심 컬럼 누락 / "
        "사업자번호 형식 / 음수·0 단가 / (정산코드, 제품명) 중복을 감지. "
        "errors > 0 이면 사용자 확인 필수."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "items": {
                "type": "array",
                "description": "cocso_settlement_create 와 동일한 items 배열",
                "items": {"type": "object"},
            },
        },
        "required": ["items"],
    },
}


# ---------------------------------------------------------------------------
# register
# ---------------------------------------------------------------------------

def _check_requirements() -> bool:
    return _OPENPYXL_AVAILABLE and TEMPLATE_PATH.exists()


def register(ctx) -> None:
    ctx.register_tool(
        name=CREATE_SCHEMA["name"],
        toolset="cocso-settlement",
        schema=CREATE_SCHEMA,
        handler=cocso_settlement_create,
        check_fn=_check_requirements,
    )
    ctx.register_tool(
        name=INFO_SCHEMA["name"],
        toolset="cocso-settlement",
        schema=INFO_SCHEMA,
        handler=cocso_settlement_template_info,
        check_fn=_check_requirements,
    )
    ctx.register_tool(
        name=VALIDATE_SCHEMA["name"],
        toolset="cocso-settlement",
        schema=VALIDATE_SCHEMA,
        handler=cocso_settlement_validate,
        check_fn=_check_requirements,
    )

    # Bundled skill — workflow guide for converting other-format Excel
    # files into the COCSO standard settlement workbook. Loaded via
    # ``skill_view("cocso_settlement:cocso-settlement-excel")``. Lives
    # under ``skills/`` inside this plugin so it disappears when the
    # plugin is disabled (no orphan docs).
    skills_dir = PLUGIN_DIR / "skills"
    if skills_dir.is_dir():
        for child in sorted(skills_dir.iterdir()):
            skill_md = child / "SKILL.md"
            if child.is_dir() and skill_md.exists():
                try:
                    ctx.register_skill(child.name, skill_md)
                except AttributeError:
                    # register_skill 미지원 런타임 (구버전) — 조용히 skip
                    logger.debug(
                        "ctx.register_skill not available; skill %s not registered",
                        child.name,
                    )
                    break
