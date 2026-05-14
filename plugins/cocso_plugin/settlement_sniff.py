"""cocso_plugin.settlement_sniff — 다양한 입력 엑셀 → COCSO 표준 매핑.

문제: 사용자가 가져오는 엑셀은 거래처마다 컬럼명·순서·헤더 행 위치가
다름. 매번 수동 매핑은 오류·반복 고통. 이 모듈은 perception layer:

  1. ``cocso_settlement_sniff`` — 자동 헤더 감지 + COCSO 표준 매칭 점수
  2. ``cocso_settlement_mapping_save`` — 사용자 확인 매핑을 preset 저장
  3. ``cocso_settlement_mapping_match`` — 새 파일이 기존 preset과 일치?
  4. ``cocso_settlement_mapping_list`` — 저장된 preset 목록

Perception 결과 (정형 items[]) 는 cocso_settlement_create 또는
mcp__cocso-service__settlement_create 가 받아서 변환·검증·xlsx 생성.
"""
from __future__ import annotations

import difflib
import hashlib
import json
import logging
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

try:
    import openpyxl  # type: ignore[import-untyped]
    _OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]
    _OPENPYXL_AVAILABLE = False

# ---------------------------------------------------------------------------
# COCSO 표준 컬럼 — settlement.py 와 일치 (single source of truth)
# (column_letter, korean_label, snake_key, type, alias_keywords)
# ---------------------------------------------------------------------------

COCSO_COLUMNS: List[Dict[str, Any]] = [
    {"col": "B", "label": "정산코드",                "key": "settlement_code",
     "type": "str", "aliases": ["정산코드", "코드", "거래코드", "code", "settlement"]},
    {"col": "C", "label": "처방 병원명",             "key": "hospital_name",
     "type": "str", "aliases": ["병원명", "처방처", "거래처", "거래처명", "의료기관", "hospital", "client"]},
    {"col": "D", "label": "처방 병원 사업자번호",    "key": "hospital_biz_id",
     "type": "str", "aliases": ["사업자번호", "사업자등록번호", "거래처사업자", "biz_id", "business_id"]},
    {"col": "E", "label": "제약사명",                "key": "manufacturer",
     "type": "str", "aliases": ["제약사", "제조사", "공급사", "공급처", "manufacturer", "supplier"]},
    {"col": "F", "label": "보험코드",                "key": "insurance_code",
     "type": "str", "aliases": ["보험코드", "보험약가코드", "약가코드", "insurance"]},
    {"col": "G", "label": "제품명",                  "key": "product_name",
     "type": "str", "aliases": ["제품명", "약품명", "의약품명", "품목명", "약품", "product", "drug"]},
    {"col": "H", "label": "단가(원)",                "key": "unit_price",
     "type": "number", "aliases": ["단가", "약가", "공급가", "단가원", "price", "unit"]},
    {"col": "I", "label": "수량",                    "key": "quantity",
     "type": "number", "aliases": ["수량", "처방수량", "청구수량", "qty", "quantity", "count"]},
    {"col": "J", "label": "처방금액(원)",            "key": "prescription_amount",
     "type": "number", "aliases": ["처방금액", "청구금액", "공급금액", "처방액", "amount", "total"]},
    {"col": "K", "label": "정산금액(원, VAT 포함)",  "key": "settlement_amount",
     "type": "number", "aliases": ["정산금액", "정산액", "정산액vat포함", "vat포함", "settlement_amount"]},
    {"col": "L", "label": "비고",                    "key": "note",
     "type": "str", "aliases": ["비고", "메모", "note", "remark", "comment"]},
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _err(msg: str, **extra) -> str:
    return json.dumps({"error": msg, **extra}, ensure_ascii=False)


def _ok(**fields) -> str:
    return json.dumps({"ok": True, **fields}, ensure_ascii=False, default=str)


def _resolve(p: str) -> Path:
    return Path(os.path.expanduser(p)).resolve()


def _check_sdk() -> Optional[str]:
    if not _OPENPYXL_AVAILABLE:
        return _err("openpyxl not installed",
                    install="pip install 'openpyxl>=3.1,<4'")
    return None


_NORMALIZE_RE = re.compile(r"[\s\(\)\[\],\.\-_/\\원]+")


def _normalize(s: Any) -> str:
    """헤더 비교용 정규화 — 공백/괄호/단위 제거, 소문자, 한글유지."""
    if s is None:
        return ""
    return _NORMALIZE_RE.sub("", str(s)).strip().lower()


def _score_match(source: str, alias: str) -> float:
    """0.0 ~ 1.0 매칭 점수. 정확>포함>fuzzy."""
    s = _normalize(source)
    a = _normalize(alias)
    if not s or not a:
        return 0.0
    if s == a:
        return 1.0
    if a in s or s in a:
        return 0.85
    return difflib.SequenceMatcher(None, s, a).ratio() * 0.8


def _best_cocso_match(source_header: str) -> Dict[str, Any]:
    """source 헤더 → COCSO 표준 컬럼 중 가장 매칭 높은 것."""
    best = {"col": None, "label": None, "key": None, "confidence": 0.0}
    for c in COCSO_COLUMNS:
        for alias in c["aliases"]:
            sc = _score_match(source_header, alias)
            if sc > best["confidence"]:
                best = {
                    "col": c["col"],
                    "label": c["label"],
                    "key": c["key"],
                    "confidence": round(sc, 3),
                    "matched_alias": alias,
                }
    return best


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------

def _detect_header_row(ws, scan_rows: int = 20) -> Optional[int]:
    """가장 가능성 높은 헤더 행 번호. 휴리스틱: 텍스트 비율 높음 + 다음
    행에 숫자 비율 높음."""
    candidates = []
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        row_vals = [ws.cell(row=r, column=c).value
                    for c in range(1, ws.max_column + 1)]
        non_empty = [v for v in row_vals if v not in (None, "")]
        if not non_empty:
            continue
        text_ratio = sum(1 for v in non_empty if isinstance(v, str)) / len(non_empty)
        next_row_vals = [ws.cell(row=r + 1, column=c).value
                         for c in range(1, ws.max_column + 1)
                         if ws.cell(row=r + 1, column=c).value not in (None, "")]
        next_num_ratio = (
            sum(1 for v in next_row_vals if isinstance(v, (int, float))) / len(next_row_vals)
            if next_row_vals else 0.0
        )
        # 헤더 score = 자기 텍스트 비율 + 다음 행 숫자 비율 가중
        score = text_ratio * 0.6 + next_num_ratio * 0.4 + min(len(non_empty), 11) / 11 * 0.2
        candidates.append((score, r, len(non_empty)))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


def _read_headers(ws, header_row: int) -> List[Dict[str, Any]]:
    """헤더 행에서 (column_letter, original_name) 추출."""
    from openpyxl.utils import get_column_letter
    out = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v in (None, ""):
            continue
        out.append({"source_col": get_column_letter(c), "source_name": str(v)})
    return out


# ---------------------------------------------------------------------------
# Tool: cocso_settlement_sniff
# ---------------------------------------------------------------------------

def cocso_settlement_sniff(args: Dict[str, Any], **_kw) -> str:
    if (e := _check_sdk()) is not None:
        return e
    path = (args.get("path") or "").strip()
    sheet_name = (args.get("sheet") or "").strip() or None
    if not path:
        return _err("path required")
    try:
        p = _resolve(path)
        if not p.exists():
            return _err(f"file not found: {p}")
        wb = openpyxl.load_workbook(p, read_only=False, data_only=True)
        sheet = sheet_name or wb.sheetnames[0]
        if sheet not in wb.sheetnames:
            wb.close()
            return _err(f"sheet '{sheet}' not found",
                        available=wb.sheetnames)
        ws = wb[sheet]
        header_row = _detect_header_row(ws)
        if header_row is None:
            wb.close()
            return _err("could not detect header row",
                        hint="시트 첫 20행 안에 텍스트 헤더가 보이지 않음")
        headers = _read_headers(ws, header_row)
        if not headers:
            wb.close()
            return _err(f"header row {header_row} is empty")

        # 각 source 헤더 → COCSO 표준 매칭
        mapped: Dict[str, Any] = {}      # cocso_label → {source_col, source_name, confidence}
        unmapped: List[Dict[str, Any]] = []
        needs_confirm: List[str] = []
        used_cocso: set = set()
        for h in headers:
            best = _best_cocso_match(h["source_name"])
            if best["confidence"] < 0.5:
                unmapped.append(h)
                continue
            label = best["label"]
            if label in used_cocso:
                # 중복 — 더 높은 confidence 가 이미 차지
                existing = mapped[label]
                if best["confidence"] > existing["confidence"]:
                    unmapped.append({"source_col": existing["source_col"],
                                     "source_name": existing["source_name"]})
                    mapped[label] = {**h, "confidence": best["confidence"],
                                     "matched_alias": best["matched_alias"]}
                else:
                    unmapped.append(h)
                continue
            used_cocso.add(label)
            mapped[label] = {**h, "confidence": best["confidence"],
                             "matched_alias": best["matched_alias"]}
            if best["confidence"] < 0.9:
                needs_confirm.append(label)

        missing = [c["label"] for c in COCSO_COLUMNS if c["label"] not in mapped]

        # 데이터 행 추정
        data_start = header_row + 1
        data_end = header_row
        for r in range(data_start, ws.max_row + 1):
            if any(ws.cell(row=r, column=c).value not in (None, "")
                   for c in range(1, ws.max_column + 1)):
                data_end = r
        data_rows_estimated = max(0, data_end - header_row)

        wb.close()
        return _ok(
            path=str(p),
            sheet=sheet,
            header_row=header_row,
            data_starts_row=data_start,
            data_rows_estimated=data_rows_estimated,
            auto_mapping=mapped,
            unmapped_source_columns=unmapped,
            missing_cocso_columns=missing,
            needs_user_confirmation=needs_confirm,
            note=(
                "낮은 confidence (< 0.9) 매핑은 사용자에게 확인 받기 권장. "
                "missing_cocso_columns는 빈 값으로 둘지 사용자 선택. "
                "결과를 cocso_settlement_mapping_save 로 저장하면 같은 양식 "
                "재방문 시 자동 적용됩니다."
            ),
        )
    except Exception as exc:
        return _err(f"sniff failed: {exc}", path=path)


# ---------------------------------------------------------------------------
# Mapping preset storage
# ---------------------------------------------------------------------------

def _mappings_dir() -> Path:
    try:
        from cocso_core.cocso_constants import get_cocso_home
        d = get_cocso_home() / "mappings"
    except Exception:
        d = Path.home() / ".cocso" / "mappings"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _signature(headers: List[str]) -> str:
    """헤더 시퀀스 정규화 → 매칭 키."""
    norm = "|".join(_normalize(h) for h in headers if h)
    return hashlib.sha1(norm.encode("utf-8")).hexdigest()[:16]


def cocso_settlement_mapping_save(args: Dict[str, Any], **_kw) -> str:
    name = (args.get("name") or "").strip()
    source_headers = args.get("source_headers") or []
    mapping = args.get("mapping") or {}  # source_name → cocso_label
    if not name or not source_headers or not mapping:
        return _err("name, source_headers (list), mapping (dict) required")
    if not isinstance(source_headers, list) or not isinstance(mapping, dict):
        return _err("source_headers must be list, mapping must be dict")
    sig = _signature(source_headers)
    fname = re.sub(r"[^\w가-힣\-]", "_", name) + f"-{sig}.json"
    path = _mappings_dir() / fname
    payload = {
        "name": name,
        "signature": sig,
        "source_headers": source_headers,
        "mapping": mapping,
        "saved_at": datetime.now(timezone.utc).isoformat(),
        "hit_count": 0,
    }
    if path.exists():
        try:
            existing = json.loads(path.read_text(encoding="utf-8"))
            payload["hit_count"] = existing.get("hit_count", 0)
        except Exception:
            pass
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return _ok(path=str(path), signature=sig, name=name)


def cocso_settlement_mapping_match(args: Dict[str, Any], **_kw) -> str:
    """현재 source headers와 일치하는 preset 찾기."""
    source_headers = args.get("source_headers") or []
    if not isinstance(source_headers, list) or not source_headers:
        return _err("source_headers (list) required")
    sig = _signature(source_headers)
    for f in _mappings_dir().glob("*.json"):
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            continue
        if data.get("signature") == sig:
            data["hit_count"] = data.get("hit_count", 0) + 1
            data["last_used"] = datetime.now(timezone.utc).isoformat()
            f.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
            return _ok(matched=True, preset=data, path=str(f))
    return _ok(matched=False, signature=sig,
               note="No preset stored for this header signature. After confirming "
                    "the mapping, call cocso_settlement_mapping_save to remember it.")


def cocso_settlement_mapping_list(args: Dict[str, Any], **_kw) -> str:
    presets = []
    for f in _mappings_dir().glob("*.json"):
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            presets.append({
                "name": data.get("name"),
                "signature": data.get("signature"),
                "saved_at": data.get("saved_at"),
                "hit_count": data.get("hit_count", 0),
                "last_used": data.get("last_used"),
                "header_count": len(data.get("source_headers", [])),
                "path": str(f),
            })
        except Exception:
            continue
    presets.sort(key=lambda x: x.get("hit_count", 0), reverse=True)
    return _ok(count=len(presets), presets=presets, dir=str(_mappings_dir()))


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------

SNIFF_SCHEMA = {
    "name": "cocso_settlement_sniff",
    "description": (
        "다양한 형식의 입력 엑셀에서 헤더 자동 감지 + COCSO 표준 11 컬럼과 "
        "fuzzy 매칭 점수 반환. 사용자가 거래처·병원 양식 엑셀을 가져왔을 때 "
        "변환 전 자동 분석 단계로 사용. 매핑 결과를 사용자에게 확인 요청 후 "
        "cocso_settlement_create / mcp__cocso-service__settlement_create 호출."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "입력 .xlsx 경로 (~ 확장 지원)"},
            "sheet": {"type": "string", "description": "시트 이름. 비우면 첫 시트."},
        },
        "required": ["path"],
    },
}

SAVE_SCHEMA = {
    "name": "cocso_settlement_mapping_save",
    "description": (
        "사용자가 확인한 매핑을 preset으로 저장. 같은 헤더 시그니처의 "
        "파일이 다시 들어오면 cocso_settlement_mapping_match 가 자동 매칭."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "name": {"type": "string", "description": "사람이 알아볼 이름 (예: '한빛약품')"},
            "source_headers": {"type": "array", "items": {"type": "string"},
                                "description": "원본 헤더 행 그대로 (순서 유지)"},
            "mapping": {"type": "object",
                         "description": "source_header → cocso_label dict"},
        },
        "required": ["name", "source_headers", "mapping"],
    },
}

MATCH_SCHEMA = {
    "name": "cocso_settlement_mapping_match",
    "description": (
        "현재 입력 파일의 source headers 와 일치하는 저장된 preset 찾기. "
        "있으면 즉시 사용 (사용자 확인 생략 가능). 없으면 sniff + 사용자 확인 흐름."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "source_headers": {"type": "array", "items": {"type": "string"}},
        },
        "required": ["source_headers"],
    },
}

LIST_SCHEMA = {
    "name": "cocso_settlement_mapping_list",
    "description": "저장된 매핑 preset 목록 — 이름 / hit_count / last_used.",
    "parameters": {"type": "object", "properties": {}},
}


# ---------------------------------------------------------------------------
# register
# ---------------------------------------------------------------------------

def _check_requirements() -> bool:
    return _OPENPYXL_AVAILABLE


def register(ctx) -> None:
    for schema, handler in (
        (SNIFF_SCHEMA, cocso_settlement_sniff),
        (SAVE_SCHEMA,  cocso_settlement_mapping_save),
        (MATCH_SCHEMA, cocso_settlement_mapping_match),
        (LIST_SCHEMA,  cocso_settlement_mapping_list),
    ):
        ctx.register_tool(
            name=schema["name"],
            toolset="cocso-settlement",
            schema=schema,
            handler=handler,
            check_fn=_check_requirements,
        )
