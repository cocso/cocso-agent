"""cocso_excel — Excel (.xlsx) 파일 읽기/쓰기 tool 모음.

6개 tool:
  excel_open         — 파일 메타 (시트 목록, 행/열 수)
  excel_read_range   — 시트 범위 (예: A1:D10) → 2D array
  excel_write_cell   — 단일 셀 값 쓰기
  excel_write_range  — 범위 batch 쓰기 (matrix)
  excel_add_sheet    — 새 시트 추가
  excel_save_as      — 다른 이름으로 저장 (원본 보호)

설계 원칙:
- 모든 핸들러는 dict 받고 JSON 문자열 반환. 예외 절대 raise 안 함.
- 쓰기 작업은 즉시 파일 저장 (xlsx는 write-on-save 모델).
- 경로는 expanduser 처리 — ``~/...`` 받음.
- 보호 경로 차단은 soul_sandbox plugin이 담당 (이 plugin은 신경 X).
- 감사 로그는 cocso_audit plugin이 자동 기록 (post_tool_call).
"""
from __future__ import annotations

import json
import logging
import os
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

try:
    import openpyxl  # type: ignore[import-untyped]
    from openpyxl.utils import get_column_letter, range_boundaries
    _OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]
    _OPENPYXL_AVAILABLE = False


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _err(msg: str, **extra) -> str:
    return json.dumps({"error": msg, **extra}, ensure_ascii=False)


def _ok(**fields) -> str:
    return json.dumps({"ok": True, **fields}, ensure_ascii=False, default=str)


def _resolve(path: str) -> Path:
    return Path(os.path.expanduser(path)).resolve()


def _check_sdk() -> Optional[str]:
    if not _OPENPYXL_AVAILABLE:
        return _err(
            "openpyxl not installed",
            install="pip install 'openpyxl>=3.1,<4'",
        )
    return None


def _open_wb(path: str, *, read_only: bool = False, data_only: bool = False):
    """Load workbook. data_only=True returns evaluated formula values."""
    p = _resolve(path)
    if not p.exists():
        raise FileNotFoundError(f"file not found: {p}")
    return openpyxl.load_workbook(
        p, read_only=read_only, data_only=data_only
    ), p


def _coerce_value(v: Any) -> Any:
    """Best-effort type coercion for cell write values from JSON."""
    if isinstance(v, str):
        # Allow leading "=" → formula; trust user.
        return v
    return v


# ---------------------------------------------------------------------------
# Tool handlers
# ---------------------------------------------------------------------------

def excel_open(args: Dict[str, Any], **_kw) -> str:
    """List sheets + size for a workbook."""
    if (e := _check_sdk()) is not None:
        return e
    path = (args.get("path") or "").strip()
    if not path:
        return _err("path required")
    try:
        wb, p = _open_wb(path, read_only=True, data_only=True)
    except Exception as exc:
        return _err(f"open failed: {exc}", path=path)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets.append({"name": name, "rows": ws.max_row, "cols": ws.max_column})
    wb.close()
    return _ok(path=str(p), sheets=sheets)


def excel_read_range(args: Dict[str, Any], **_kw) -> str:
    """Read a rectangular range. Returns 2D array of values."""
    if (e := _check_sdk()) is not None:
        return e
    path = (args.get("path") or "").strip()
    sheet = (args.get("sheet") or "").strip()
    rng = (args.get("range") or "").strip()
    if not path or not sheet or not rng:
        return _err("path, sheet, range required")
    try:
        wb, _p = _open_wb(path, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return _err(f"sheet '{sheet}' not found",
                        available=wb.sheetnames if wb else [])
        ws = wb[sheet]
        try:
            min_col, min_row, max_col, max_row = range_boundaries(rng)
        except Exception as exc:
            wb.close()
            return _err(f"invalid range '{rng}': {exc}")
        rows: List[List[Any]] = []
        for row in ws.iter_rows(
            min_row=min_row, max_row=max_row,
            min_col=min_col, max_col=max_col, values_only=True,
        ):
            rows.append(list(row))
        wb.close()
        return _ok(sheet=sheet, range=rng, rows=rows,
                   shape=[len(rows), len(rows[0]) if rows else 0])
    except Exception as exc:
        return _err(f"read failed: {exc}", path=path, sheet=sheet, range=rng)


def excel_write_cell(args: Dict[str, Any], **_kw) -> str:
    """Write a single cell. Saves immediately."""
    if (e := _check_sdk()) is not None:
        return e
    path = (args.get("path") or "").strip()
    sheet = (args.get("sheet") or "").strip()
    cell = (args.get("cell") or "").strip()
    if "value" not in args:
        return _err("value required (use null to clear)")
    if not path or not sheet or not cell:
        return _err("path, sheet, cell required")
    try:
        wb, p = _open_wb(path)
        if sheet not in wb.sheetnames:
            wb.close()
            return _err(f"sheet '{sheet}' not found",
                        available=wb.sheetnames)
        ws = wb[sheet]
        ws[cell] = _coerce_value(args["value"])
        wb.save(p)
        wb.close()
        return _ok(path=str(p), sheet=sheet, cell=cell, value=args["value"])
    except Exception as exc:
        return _err(f"write failed: {exc}", path=path, sheet=sheet, cell=cell)


def excel_write_range(args: Dict[str, Any], **_kw) -> str:
    """Write a 2D matrix starting at a top-left cell. Saves immediately."""
    if (e := _check_sdk()) is not None:
        return e
    path = (args.get("path") or "").strip()
    sheet = (args.get("sheet") or "").strip()
    start = (args.get("start_cell") or "").strip()
    values = args.get("values")
    if not path or not sheet or not start:
        return _err("path, sheet, start_cell required")
    if not isinstance(values, list) or not all(isinstance(r, list) for r in values):
        return _err("values must be a 2D array (list of lists)")
    try:
        wb, p = _open_wb(path)
        if sheet not in wb.sheetnames:
            wb.close()
            return _err(f"sheet '{sheet}' not found",
                        available=wb.sheetnames)
        ws = wb[sheet]
        try:
            start_col, start_row, _, _ = range_boundaries(start)
        except Exception as exc:
            wb.close()
            return _err(f"invalid start_cell '{start}': {exc}")
        rows_written = 0
        cols_written = 0
        for r_offset, row_vals in enumerate(values):
            for c_offset, v in enumerate(row_vals):
                ws.cell(
                    row=start_row + r_offset,
                    column=start_col + c_offset,
                    value=_coerce_value(v),
                )
            rows_written = max(rows_written, r_offset + 1)
            cols_written = max(cols_written, len(row_vals))
        wb.save(p)
        wb.close()
        end_cell = f"{get_column_letter(start_col + cols_written - 1)}{start_row + rows_written - 1}"
        return _ok(path=str(p), sheet=sheet, range=f"{start}:{end_cell}",
                   shape=[rows_written, cols_written])
    except Exception as exc:
        return _err(f"write failed: {exc}", path=path, sheet=sheet, start_cell=start)


def excel_add_sheet(args: Dict[str, Any], **_kw) -> str:
    """Create a new sheet. Saves immediately."""
    if (e := _check_sdk()) is not None:
        return e
    path = (args.get("path") or "").strip()
    name = (args.get("name") or "").strip()
    if not path or not name:
        return _err("path, name required")
    try:
        wb, p = _open_wb(path)
        if name in wb.sheetnames:
            wb.close()
            return _err(f"sheet '{name}' already exists",
                        available=wb.sheetnames)
        wb.create_sheet(name)
        wb.save(p)
        wb.close()
        return _ok(path=str(p), sheet=name, sheets_after=wb.sheetnames)
    except Exception as exc:
        return _err(f"add_sheet failed: {exc}", path=path, name=name)


def excel_save_as(args: Dict[str, Any], **_kw) -> str:
    """Open ``path`` and save a copy at ``dest_path``. Original untouched."""
    if (e := _check_sdk()) is not None:
        return e
    path = (args.get("path") or "").strip()
    dest = (args.get("dest_path") or "").strip()
    if not path or not dest:
        return _err("path, dest_path required")
    try:
        src = _resolve(path)
        dst = _resolve(dest)
        if not src.exists():
            return _err(f"source not found: {src}")
        wb = openpyxl.load_workbook(src)
        dst.parent.mkdir(parents=True, exist_ok=True)
        wb.save(dst)
        wb.close()
        return _ok(path=str(src), dest_path=str(dst))
    except Exception as exc:
        return _err(f"save_as failed: {exc}", path=path, dest_path=dest)


# ---------------------------------------------------------------------------
# Tool schemas
# ---------------------------------------------------------------------------

OPEN_SCHEMA = {
    "name": "excel_open",
    "description": (
        "Open an Excel (.xlsx) file and return its sheet list with size. "
        "Use this first to discover the structure before reading or writing."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Path to .xlsx file (~ expanded)"},
        },
        "required": ["path"],
    },
}

READ_RANGE_SCHEMA = {
    "name": "excel_read_range",
    "description": (
        "Read a rectangular range from a sheet. Range is A1-style "
        "(e.g. 'A1:D10'). Returns evaluated formula values."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "sheet": {"type": "string"},
            "range": {"type": "string", "description": "A1-style (e.g. 'A1:D10')"},
        },
        "required": ["path", "sheet", "range"],
    },
}

WRITE_CELL_SCHEMA = {
    "name": "excel_write_cell",
    "description": (
        "Write a single cell value. Saves immediately. Strings starting "
        "with '=' are treated as formulas. Use null to clear."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "sheet": {"type": "string"},
            "cell": {"type": "string", "description": "A1-style (e.g. 'B5')"},
            "value": {
                "description": "Cell value (string, number, bool, null). String starting with '=' is a formula.",
            },
        },
        "required": ["path", "sheet", "cell", "value"],
    },
}

WRITE_RANGE_SCHEMA = {
    "name": "excel_write_range",
    "description": (
        "Batch-write a 2D matrix starting at a top-left cell. Saves "
        "immediately. Use for inserting tables / multiple rows at once."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "sheet": {"type": "string"},
            "start_cell": {"type": "string", "description": "Top-left A1-style cell"},
            "values": {
                "type": "array",
                "description": "2D matrix: array of arrays of cell values",
                "items": {"type": "array"},
            },
        },
        "required": ["path", "sheet", "start_cell", "values"],
    },
}

ADD_SHEET_SCHEMA = {
    "name": "excel_add_sheet",
    "description": "Create a new empty sheet in the workbook. Saves immediately.",
    "parameters": {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "name": {"type": "string", "description": "New sheet name (must not already exist)"},
        },
        "required": ["path", "name"],
    },
}

SAVE_AS_SCHEMA = {
    "name": "excel_save_as",
    "description": (
        "Open a workbook and save a full copy at a new path. The original "
        "is not modified. Useful for snapshotting before edits."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "dest_path": {"type": "string"},
        },
        "required": ["path", "dest_path"],
    },
}


# ---------------------------------------------------------------------------
# register
# ---------------------------------------------------------------------------

def _check_requirements() -> bool:
    return _OPENPYXL_AVAILABLE


def register(ctx) -> None:
    for schema, handler in (
        (OPEN_SCHEMA,        excel_open),
        (READ_RANGE_SCHEMA,  excel_read_range),
        (WRITE_CELL_SCHEMA,  excel_write_cell),
        (WRITE_RANGE_SCHEMA, excel_write_range),
        (ADD_SHEET_SCHEMA,   excel_add_sheet),
        (SAVE_AS_SCHEMA,     excel_save_as),
    ):
        ctx.register_tool(
            name=schema["name"],
            toolset="excel",
            schema=schema,
            handler=handler,
            check_fn=_check_requirements,
        )
