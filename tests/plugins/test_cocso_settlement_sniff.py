"""Tests for ``cocso_settlement_sniff`` and mapping preset tools.

Pins:
- Header row auto-detected from sample 거래처 file
- 10/11 COCSO standard columns auto-mapped (보험코드 missing)
- needs_user_confirmation 비어있어야 함 (모두 high confidence)
- mapping save → match roundtrip 동작
- mapping_list 정렬 (hit_count DESC)
- 핸들러는 raise 안 함, JSON 반환
- COCSO_HOME 격리 — 테스트 간 mapping preset leak 없음
"""
from __future__ import annotations

import importlib.util
import json
import os
from pathlib import Path

import openpyxl
import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
PLUGIN_PATH = PROJECT_ROOT / "plugins" / "cocso_plugin" / "settlement_sniff.py"
SAMPLE_PATH = PROJECT_ROOT / "excel" / "samples" / "거래처_원본_2026-05.xlsx"


@pytest.fixture
def sniff(monkeypatch, tmp_path):
    """Fresh module + isolated COCSO_HOME for mapping preset storage."""
    home = tmp_path / "cocso_home"
    home.mkdir()
    monkeypatch.setenv("COCSO_HOME", str(home))

    spec = importlib.util.spec_from_file_location(
        f"_sniff_test_{tmp_path.name}", str(PLUGIN_PATH)
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod, home


def _decode(s):
    return json.loads(s)


# ---------------------------------------------------------------------------
# Sample-driven sniff
# ---------------------------------------------------------------------------

class TestSniffSample:
    def test_detects_header_row(self, sniff):
        mod, _ = sniff
        out = _decode(mod.cocso_settlement_sniff({"path": str(SAMPLE_PATH)}))
        assert out["ok"] is True
        assert out["header_row"] == 4
        assert out["data_starts_row"] == 5

    def test_estimates_data_rows(self, sniff):
        mod, _ = sniff
        out = _decode(mod.cocso_settlement_sniff({"path": str(SAMPLE_PATH)}))
        assert out["data_rows_estimated"] == 8

    def test_auto_maps_ten_of_eleven(self, sniff):
        mod, _ = sniff
        out = _decode(mod.cocso_settlement_sniff({"path": str(SAMPLE_PATH)}))
        mapped = set(out["auto_mapping"].keys())
        # 보험코드만 missing — 나머지 10개 매핑되어야
        assert len(mapped) == 10
        assert "보험코드" not in mapped

    def test_high_confidence_no_user_confirmation_needed(self, sniff):
        mod, _ = sniff
        out = _decode(mod.cocso_settlement_sniff({"path": str(SAMPLE_PATH)}))
        # sample은 흔한 패턴 — confidence 모두 0.9 이상
        assert out["needs_user_confirmation"] == []

    def test_missing_columns_listed(self, sniff):
        mod, _ = sniff
        out = _decode(mod.cocso_settlement_sniff({"path": str(SAMPLE_PATH)}))
        assert out["missing_cocso_columns"] == ["보험코드"]


class TestSniffErrors:
    def test_missing_path(self, sniff):
        mod, _ = sniff
        out = _decode(mod.cocso_settlement_sniff({}))
        assert "error" in out

    def test_file_not_found(self, sniff):
        mod, _ = sniff
        out = _decode(mod.cocso_settlement_sniff({"path": "/no/such/file.xlsx"}))
        assert "error" in out

    def test_unknown_sheet(self, sniff):
        mod, _ = sniff
        out = _decode(mod.cocso_settlement_sniff(
            {"path": str(SAMPLE_PATH), "sheet": "NoSuchSheet"}
        ))
        assert "error" in out

    def test_handler_never_raises(self, sniff):
        mod, _ = sniff
        for bad in ({}, {"path": None}, {"path": "", "sheet": None}):
            assert isinstance(mod.cocso_settlement_sniff(bad), str)


class TestHeaderHeuristic:
    def test_picks_text_heavy_row_above_numeric_data(self, sniff, tmp_path):
        """Custom workbook: row 1 = title, row 3 = headers, row 4+ = numeric data."""
        mod, _ = sniff
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "정산서 v2"  # title — text but next row not numeric
        ws["A2"] = ""  # blank
        # row 3 = headers
        ws["A3"], ws["B3"], ws["C3"] = "정산코드", "처방 병원명", "단가"
        # row 4-6 = data (numeric/text mix in correct slots)
        ws["A4"], ws["B4"], ws["C4"] = "S1", "병원A", 100
        ws["A5"], ws["B5"], ws["C5"] = "S2", "병원B", 200
        path = tmp_path / "custom.xlsx"
        wb.save(path)
        wb.close()

        out = _decode(mod.cocso_settlement_sniff({"path": str(path)}))
        assert out["ok"] is True
        # heuristic should pick row 3 since row below has numeric ratio
        assert out["header_row"] == 3


# ---------------------------------------------------------------------------
# Mapping preset roundtrip
# ---------------------------------------------------------------------------

class TestMappingSaveLoad:
    def test_save_then_match(self, sniff):
        mod, home = sniff
        headers = ["거래코드", "거래처", "사업자번호", "약품명", "약가", "수량"]
        mapping = {"거래코드": "정산코드", "거래처": "처방 병원명",
                   "사업자번호": "처방 병원 사업자번호", "약품명": "제품명",
                   "약가": "단가(원)", "수량": "수량"}
        saved = _decode(mod.cocso_settlement_mapping_save({
            "name": "한빛약품", "source_headers": headers, "mapping": mapping,
        }))
        assert saved["ok"] is True
        assert saved["name"] == "한빛약품"

        matched = _decode(mod.cocso_settlement_mapping_match(
            {"source_headers": headers}
        ))
        assert matched["matched"] is True
        assert matched["preset"]["name"] == "한빛약품"

    def test_unknown_signature_returns_no_match(self, sniff):
        mod, _ = sniff
        out = _decode(mod.cocso_settlement_mapping_match(
            {"source_headers": ["완전히다른헤더1", "다른2"]}
        ))
        assert out["matched"] is False
        assert "signature" in out

    def test_match_increments_hit_count(self, sniff):
        mod, _ = sniff
        headers = ["a", "b", "c"]
        mod.cocso_settlement_mapping_save({
            "name": "x", "source_headers": headers, "mapping": {"a": "정산코드"}
        })
        for _ in range(3):
            mod.cocso_settlement_mapping_match({"source_headers": headers})
        listed = _decode(mod.cocso_settlement_mapping_list({}))
        assert listed["count"] == 1
        assert listed["presets"][0]["hit_count"] == 3

    def test_save_validates_args(self, sniff):
        mod, _ = sniff
        out = _decode(mod.cocso_settlement_mapping_save({
            "name": "x", "source_headers": "not-list", "mapping": {}
        }))
        assert "error" in out

    def test_list_sorted_by_hit_count(self, sniff):
        mod, _ = sniff
        # Save 3 presets, hit different number of times
        for name, headers, hits in (
            ("a", ["h1", "h2"], 1),
            ("b", ["x1", "x2"], 5),
            ("c", ["y1", "y2"], 2),
        ):
            mod.cocso_settlement_mapping_save({
                "name": name, "source_headers": headers, "mapping": {"h": "정산코드"}
            })
            for _ in range(hits):
                mod.cocso_settlement_mapping_match({"source_headers": headers})
        listed = _decode(mod.cocso_settlement_mapping_list({}))
        names_order = [p["name"] for p in listed["presets"]]
        assert names_order == ["b", "c", "a"]  # 5, 2, 1


# ---------------------------------------------------------------------------
# Score / normalization
# ---------------------------------------------------------------------------

class TestScoring:
    def test_normalize_strips_units_and_punctuation(self, sniff):
        mod, _ = sniff
        assert mod._normalize("단가(원)") == mod._normalize("단가")
        assert mod._normalize("정산금액(원, VAT 포함)") == mod._normalize("정산금액vat포함")

    def test_exact_match_score_is_one(self, sniff):
        mod, _ = sniff
        assert mod._score_match("정산코드", "정산코드") == 1.0

    def test_substring_match_high(self, sniff):
        mod, _ = sniff
        assert mod._score_match("거래처명", "거래처") > 0.7


# ---------------------------------------------------------------------------
# register
# ---------------------------------------------------------------------------

class TestRegister:
    def test_register_attaches_four_tools(self, sniff):
        mod, _ = sniff

        class FakeCtx:
            def __init__(self):
                self.tools = []

            def register_tool(self, **kw):
                self.tools.append(kw["name"])

        ctx = FakeCtx()
        mod.register(ctx)
        assert set(ctx.tools) == {
            "cocso_settlement_sniff",
            "cocso_settlement_mapping_save",
            "cocso_settlement_mapping_match",
            "cocso_settlement_mapping_list",
        }
