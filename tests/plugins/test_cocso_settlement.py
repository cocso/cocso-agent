"""Tests for the cocso_settlement plugin.

Pins:
- Template is bundled and discoverable
- create() copies template, fills L2 + data rows from row 4
- Both Korean labels and snake_case keys map to the same columns
- Unknown keys do not crash; they are surfaced via unknown_keys
- Numeric coercion handles "1,200" style strings
- Overwrite refused unless explicit
- Template capacity respected
- Statistics region (N3:R) untouched after our writes — template formulas remain
"""
from __future__ import annotations

import importlib.util
import json
from pathlib import Path

import openpyxl
import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
PLUGIN_PATH = PROJECT_ROOT / "plugins" / "cocso_plugin" / "settlement.py"


@pytest.fixture
def settlement(tmp_path):
    spec = importlib.util.spec_from_file_location(
        f"_settle_test_{tmp_path.name}", str(PLUGIN_PATH)
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod, tmp_path


def _decode(s: str) -> dict:
    return json.loads(s)


class TestTemplate:
    def test_template_bundled(self, settlement):
        mod, _ = settlement
        assert mod.TEMPLATE_PATH.exists()
        assert mod.TEMPLATE_PATH.suffix == ".xlsx"

    def test_template_info(self, settlement):
        mod, _ = settlement
        out = _decode(mod.cocso_settlement_template_info({}))
        assert out["ok"] is True
        assert out["sheet_name"] == "의약품 정산 수수료 내역서"
        assert out["company_cell"] == "L2"
        assert out["data_start_row"] == 4
        assert len(out["columns"]) == 11
        # spot-check first + last column
        first = out["columns"][0]
        assert first["column"] == "B" and first["snake_key"] == "settlement_code"
        last = out["columns"][-1]
        assert last["column"] == "L" and last["snake_key"] == "note"


class TestCreateBasics:
    def test_writes_file(self, settlement):
        mod, tmp = settlement
        dst = tmp / "정산.xlsx"
        out = _decode(mod.cocso_settlement_create({
            "output_path": str(dst),
            "company_name": "테스트약품",
            "items": [],
        }))
        assert out["ok"] is True
        assert dst.exists()

    def test_company_name_lands_in_L2(self, settlement):
        mod, tmp = settlement
        dst = tmp / "정산.xlsx"
        mod.cocso_settlement_create({
            "output_path": str(dst),
            "company_name": "OO약품",
            "items": [],
        })
        wb = openpyxl.load_workbook(dst)
        ws = wb["의약품 정산 수수료 내역서"]
        assert ws["L2"].value == "OO약품"
        wb.close()

    def test_data_starts_at_row_4(self, settlement):
        mod, tmp = settlement
        dst = tmp / "정산.xlsx"
        mod.cocso_settlement_create({
            "output_path": str(dst),
            "company_name": "OO",
            "items": [
                {"settlement_code": "S0001", "hospital_name": "가나다병원",
                 "unit_price": 1000, "quantity": 5},
            ],
        })
        wb = openpyxl.load_workbook(dst)
        ws = wb["의약품 정산 수수료 내역서"]
        assert ws["B4"].value == "S0001"
        assert ws["C4"].value == "가나다병원"
        assert ws["H4"].value == 1000
        assert ws["I4"].value == 5
        wb.close()


class TestKeyMappings:
    def test_korean_keys(self, settlement):
        mod, tmp = settlement
        dst = tmp / "정산.xlsx"
        out = _decode(mod.cocso_settlement_create({
            "output_path": str(dst),
            "company_name": "OO",
            "items": [{
                "정산코드": "S0001",
                "처방 병원명": "한빛의원",
                "처방 병원 사업자번호": "123-45-67890",
                "제약사명": "코쏘제약",
                "보험코드": "INS-001",
                "제품명": "타이레놀",
                "단가(원)": 500,
                "수량": 100,
                "처방금액(원)": 50000,
                "정산금액(원, VAT 포함)": 55000,
                "비고": "월말 정산",
            }],
        }))
        assert out["ok"] is True
        assert out["rows_written"] == 1
        assert out.get("unknown_keys") is None
        wb = openpyxl.load_workbook(dst)
        ws = wb["의약품 정산 수수료 내역서"]
        assert ws["B4"].value == "S0001"
        assert ws["C4"].value == "한빛의원"
        assert ws["L4"].value == "월말 정산"
        wb.close()

    def test_snake_keys(self, settlement):
        mod, tmp = settlement
        dst = tmp / "정산.xlsx"
        out = _decode(mod.cocso_settlement_create({
            "output_path": str(dst),
            "company_name": "OO",
            "items": [{
                "settlement_code": "S0002",
                "hospital_name": "두번째병원",
                "manufacturer": "두번째제약",
                "product_name": "약품B",
                "unit_price": 2000,
                "quantity": 3,
            }],
        }))
        assert out["ok"] is True
        wb = openpyxl.load_workbook(dst)
        ws = wb["의약품 정산 수수료 내역서"]
        assert ws["B4"].value == "S0002"
        assert ws["E4"].value == "두번째제약"
        wb.close()

    def test_unknown_keys_surfaced(self, settlement):
        mod, tmp = settlement
        dst = tmp / "정산.xlsx"
        out = _decode(mod.cocso_settlement_create({
            "output_path": str(dst),
            "company_name": "OO",
            "items": [{
                "오타필드": "값",
                "another_unknown": 123,
            }],
        }))
        assert out["ok"] is True
        # Item에 인식 키 없음 → row 안 채움 → unknown_keys 표면
        assert out["rows_written"] == 0
        assert "오타필드" in out["unknown_keys"]
        assert "another_unknown" in out["unknown_keys"]


class TestNumericCoercion:
    def test_comma_string_becomes_int(self, settlement):
        mod, tmp = settlement
        dst = tmp / "정산.xlsx"
        mod.cocso_settlement_create({
            "output_path": str(dst), "company_name": "OO",
            "items": [{"unit_price": "1,200,000", "quantity": "5"}],
        })
        wb = openpyxl.load_workbook(dst)
        ws = wb["의약품 정산 수수료 내역서"]
        assert ws["H4"].value == 1200000
        assert ws["I4"].value == 5
        wb.close()

    def test_string_passthrough_when_uncoercible(self, settlement):
        mod, tmp = settlement
        dst = tmp / "정산.xlsx"
        mod.cocso_settlement_create({
            "output_path": str(dst), "company_name": "OO",
            "items": [{"unit_price": "약 1만원"}],
        })
        wb = openpyxl.load_workbook(dst)
        ws = wb["의약품 정산 수수료 내역서"]
        # leaves the original string in place
        assert ws["H4"].value == "약 1만원"
        wb.close()


class TestOverwriteGuard:
    def test_refuses_existing_file(self, settlement):
        mod, tmp = settlement
        dst = tmp / "정산.xlsx"
        dst.write_text("existing")
        out = _decode(mod.cocso_settlement_create({
            "output_path": str(dst), "company_name": "OO", "items": [],
        }))
        assert "error" in out
        assert "exists" in out["error"].lower()

    def test_overwrite_flag_replaces(self, settlement):
        mod, tmp = settlement
        dst = tmp / "정산.xlsx"
        dst.write_text("existing")
        out = _decode(mod.cocso_settlement_create({
            "output_path": str(dst), "company_name": "OO", "items": [],
            "overwrite": True,
        }))
        assert out["ok"] is True


class TestCapacityAndPath:
    def test_rejects_non_xlsx(self, settlement):
        mod, tmp = settlement
        out = _decode(mod.cocso_settlement_create({
            "output_path": str(tmp / "정산.csv"),
            "company_name": "OO", "items": [],
        }))
        assert "error" in out

    def test_rejects_too_many_items(self, settlement):
        mod, tmp = settlement
        too_many = [{"settlement_code": f"S{i:04d}"} for i in range(5000)]
        out = _decode(mod.cocso_settlement_create({
            "output_path": str(tmp / "정산.xlsx"),
            "company_name": "OO", "items": too_many,
        }))
        assert "error" in out
        assert "exceed" in out["error"].lower() or "capacity" in out["error"].lower()


class TestStatisticsRegionIntact:
    def test_template_formulas_preserved(self, settlement):
        """N3:R cells with formulas must survive a create() call."""
        mod, tmp = settlement
        dst = tmp / "정산.xlsx"
        mod.cocso_settlement_create({
            "output_path": str(dst), "company_name": "OO",
            "items": [{"settlement_code": "S0001", "unit_price": 100, "quantity": 2}],
        })
        wb = openpyxl.load_workbook(dst)
        ws = wb["의약품 정산 수수료 내역서"]
        # The template has a SUM formula at O4 over J4:J5000.
        cell = ws["O4"]
        # value can be the formula string OR the cached value depending on
        # openpyxl behavior; just ensure it's not None / blank.
        assert cell.value is not None
        wb.close()


class TestRegister:
    def test_register_attaches_three_tools(self, settlement):
        mod, _ = settlement

        class FakeCtx:
            def __init__(self):
                self.tools = []

            def register_tool(self, **kw):
                self.tools.append(kw["name"])

        ctx = FakeCtx()
        mod.register(ctx)
        assert set(ctx.tools) == {
            "cocso_settlement_create",
            "cocso_settlement_template_info",
            "cocso_settlement_validate",
        }
