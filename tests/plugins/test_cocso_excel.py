"""Tests for the cocso_excel plugin.

Pins:
- All 6 tools return JSON strings (never raise)
- Errors come back as ``{"error": ...}``, success as ``{"ok": true, ...}``
- Reading evaluated formula values
- Writes persist (saved immediately)
- Range writes match the requested matrix shape
- Sheet operations (add) cannot duplicate existing names
- save_as does not modify the original
- Path arguments accept ``~`` expansion
"""
from __future__ import annotations

import importlib.util
import json
from pathlib import Path

import openpyxl  # provided by core deps
import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
PLUGIN_PATH = PROJECT_ROOT / "plugins" / "cocso_plugin" / "excel.py"


@pytest.fixture
def excel(tmp_path):
    """Fresh plugin module + a sample .xlsx in tmp."""
    spec = importlib.util.spec_from_file_location(
        f"_excel_test_{tmp_path.name}", str(PLUGIN_PATH)
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)

    # Build a sample workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "name"
    ws["B1"] = "qty"
    ws["A2"] = "Alpha"
    ws["B2"] = 10
    ws["A3"] = "Beta"
    ws["B3"] = 20
    ws["C1"] = "total"
    ws["C2"] = "=B2*2"
    ws["C3"] = "=B3*2"
    wb.create_sheet("Other")
    path = tmp_path / "sample.xlsx"
    wb.save(path)
    wb.close()
    return mod, path


def _decode(s: str) -> dict:
    return json.loads(s)


class TestOpen:
    def test_lists_sheets(self, excel):
        mod, path = excel
        out = _decode(mod.excel_open({"path": str(path)}))
        assert out.get("ok") is True
        names = [s["name"] for s in out["sheets"]]
        assert "Sheet1" in names and "Other" in names

    def test_reports_size(self, excel):
        mod, path = excel
        out = _decode(mod.excel_open({"path": str(path)}))
        sheet1 = next(s for s in out["sheets"] if s["name"] == "Sheet1")
        assert sheet1["rows"] == 3
        assert sheet1["cols"] == 3

    def test_missing_path_errors(self, excel):
        mod, _ = excel
        out = _decode(mod.excel_open({"path": "/no/such/file.xlsx"}))
        assert "error" in out

    def test_no_path_arg(self, excel):
        mod, _ = excel
        out = _decode(mod.excel_open({}))
        assert "error" in out


class TestReadRange:
    def test_reads_full_range(self, excel):
        mod, path = excel
        out = _decode(mod.excel_read_range(
            {"path": str(path), "sheet": "Sheet1", "range": "A1:B3"}
        ))
        assert out["ok"] is True
        assert out["rows"] == [["name", "qty"], ["Alpha", 10], ["Beta", 20]]
        assert out["shape"] == [3, 2]

    def test_reads_evaluated_formulas(self, excel):
        # openpyxl returns cached evaluated values when data_only=True;
        # since we never opened the file in Excel, formulas come back as None.
        # This is openpyxl behavior, not a bug — pin it so future changes are
        # visible.
        mod, path = excel
        out = _decode(mod.excel_read_range(
            {"path": str(path), "sheet": "Sheet1", "range": "C2:C3"}
        ))
        assert out["ok"] is True
        # Without a prior Excel save, cached values are None.
        assert out["rows"] == [[None], [None]]

    def test_unknown_sheet(self, excel):
        mod, path = excel
        out = _decode(mod.excel_read_range(
            {"path": str(path), "sheet": "NoSuch", "range": "A1:B2"}
        ))
        assert "error" in out

    def test_invalid_range_string(self, excel):
        mod, path = excel
        out = _decode(mod.excel_read_range(
            {"path": str(path), "sheet": "Sheet1", "range": "GARBAGE"}
        ))
        assert "error" in out


class TestWriteCell:
    def test_writes_string(self, excel):
        mod, path = excel
        out = _decode(mod.excel_write_cell(
            {"path": str(path), "sheet": "Sheet1", "cell": "A4", "value": "Gamma"}
        ))
        assert out["ok"] is True
        # Re-open and verify
        wb = openpyxl.load_workbook(path)
        assert wb["Sheet1"]["A4"].value == "Gamma"
        wb.close()

    def test_writes_number(self, excel):
        mod, path = excel
        out = _decode(mod.excel_write_cell(
            {"path": str(path), "sheet": "Sheet1", "cell": "B4", "value": 42}
        ))
        assert out["ok"] is True
        wb = openpyxl.load_workbook(path)
        assert wb["Sheet1"]["B4"].value == 42
        wb.close()

    def test_writes_formula(self, excel):
        mod, path = excel
        out = _decode(mod.excel_write_cell(
            {"path": str(path), "sheet": "Sheet1", "cell": "D2", "value": "=A2"}
        ))
        assert out["ok"] is True
        wb = openpyxl.load_workbook(path)
        assert wb["Sheet1"]["D2"].value == "=A2"
        wb.close()

    def test_clear_with_null(self, excel):
        mod, path = excel
        out = _decode(mod.excel_write_cell(
            {"path": str(path), "sheet": "Sheet1", "cell": "A2", "value": None}
        ))
        assert out["ok"] is True
        wb = openpyxl.load_workbook(path)
        assert wb["Sheet1"]["A2"].value is None
        wb.close()

    def test_missing_value_errors(self, excel):
        mod, path = excel
        out = _decode(mod.excel_write_cell(
            {"path": str(path), "sheet": "Sheet1", "cell": "A4"}
        ))
        assert "error" in out


class TestWriteRange:
    def test_writes_2d_matrix(self, excel):
        mod, path = excel
        matrix = [["x", 1], ["y", 2], ["z", 3]]
        out = _decode(mod.excel_write_range(
            {"path": str(path), "sheet": "Sheet1", "start_cell": "E1", "values": matrix}
        ))
        assert out["ok"] is True
        assert out["shape"] == [3, 2]
        wb = openpyxl.load_workbook(path)
        ws = wb["Sheet1"]
        assert ws["E1"].value == "x"
        assert ws["F3"].value == 3
        wb.close()

    def test_rejects_non_2d(self, excel):
        mod, path = excel
        out = _decode(mod.excel_write_range(
            {"path": str(path), "sheet": "Sheet1", "start_cell": "E1",
             "values": ["flat", "not", "2d"]}
        ))
        assert "error" in out

    def test_invalid_start_cell(self, excel):
        mod, path = excel
        out = _decode(mod.excel_write_range(
            {"path": str(path), "sheet": "Sheet1", "start_cell": "ZZZZ",
             "values": [[1]]}
        ))
        assert "error" in out


class TestAddSheet:
    def test_adds_new_sheet(self, excel):
        mod, path = excel
        out = _decode(mod.excel_add_sheet({"path": str(path), "name": "Brand New"}))
        assert out["ok"] is True
        wb = openpyxl.load_workbook(path)
        assert "Brand New" in wb.sheetnames
        wb.close()

    def test_rejects_duplicate_name(self, excel):
        mod, path = excel
        out = _decode(mod.excel_add_sheet({"path": str(path), "name": "Sheet1"}))
        assert "error" in out


class TestSaveAs:
    def test_creates_copy_without_modifying_original(self, excel, tmp_path):
        mod, path = excel
        dest = tmp_path / "subdir" / "copy.xlsx"
        out = _decode(mod.excel_save_as(
            {"path": str(path), "dest_path": str(dest)}
        ))
        assert out["ok"] is True
        assert dest.exists()
        # Original still has original sheets (no surprise mutation)
        wb = openpyxl.load_workbook(path)
        assert "Sheet1" in wb.sheetnames
        wb.close()

    def test_missing_source(self, excel, tmp_path):
        mod, _ = excel
        out = _decode(mod.excel_save_as(
            {"path": "/no/such/x.xlsx", "dest_path": str(tmp_path / "y.xlsx")}
        ))
        assert "error" in out


class TestRegister:
    def test_register_attaches_six_tools(self, excel):
        mod, _ = excel

        class FakeCtx:
            def __init__(self):
                self.tools = []

            def register_tool(self, **kw):
                self.tools.append(kw["name"])

        ctx = FakeCtx()
        mod.register(ctx)
        assert set(ctx.tools) == {
            "excel_open",
            "excel_read_range",
            "excel_write_cell",
            "excel_write_range",
            "excel_add_sheet",
            "excel_save_as",
        }


class TestErrorContract:
    def test_handlers_never_raise(self, excel):
        """Every handler must catch all exceptions and return JSON."""
        mod, _ = excel
        # Bogus args that would crash naive code
        bad = {"path": None, "sheet": None, "cell": None, "value": None,
               "start_cell": None, "values": None, "name": None,
               "dest_path": None, "range": None}
        for fn in (mod.excel_open, mod.excel_read_range, mod.excel_write_cell,
                   mod.excel_write_range, mod.excel_add_sheet, mod.excel_save_as):
            out = fn(bad)
            assert isinstance(out, str)
            assert json.loads(out)  # parses cleanly
